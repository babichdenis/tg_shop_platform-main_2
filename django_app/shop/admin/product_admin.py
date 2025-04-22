# django_app/shop/admin/product_admin.py
import json
import csv
import os
from io import StringIO, TextIOWrapper
from django import forms
from django.urls import path
from django.contrib import admin, messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db import transaction
from django.core.files.storage import default_storage
import openpyxl
import logging
from .base import BaseAdmin
from ..models import Product, Category
from ..forms import ProductExportForm

logger = logging.getLogger(__name__)

class ImportProductsForm(forms.Form):
    FILE_FORMAT_CHOICES = [
        ('csv', 'CSV'),
        ('json', 'JSON'),
        ('xlsx', 'Excel (XLSX)'),
    ]
    file_format = forms.ChoiceField(choices=FILE_FORMAT_CHOICES, label="Формат файла")
    file = forms.FileField(label="Файл")

    def clean_file(self):
        file = self.cleaned_data['file']
        file_format = self.cleaned_data.get('file_format')
        valid_extensions = {
            'csv': ['.csv'],
            'json': ['.json'],
            'xlsx': ['.xlsx'],
        }
        extension = os.path.splitext(file.name)[1].lower()
        if extension not in valid_extensions.get(file_format, []):
            raise forms.ValidationError(
                f"Недопустимое расширение файла для формата {file_format}. Ожидается: {', '.join(valid_extensions[file_format])}"
            )
        return file

@admin.register(Product)
class ProductAdmin(BaseAdmin):
    list_display = ('id', 'name_colored', 'category', 'price', 'created_at', 'is_active')
    search_fields = ('name', 'description')
    list_filter = ('category', 'is_active')
    actions = BaseAdmin.actions + ['export_products']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import/', self.admin_site.admin_view(self.import_products_view), name='import_products'),
        ]
        return custom_urls + urls

    def import_products_view(self, request):
        if request.method == 'POST':
            form = ImportProductsForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']
                file_format = form.cleaned_data['file_format']
                errors = []
                imported_count = 0

                try:
                    with transaction.atomic():
                        if file_format == 'json':
                            data = self.parse_json(file)
                        elif file_format == 'csv':
                            data = self.csv_to_json(file)
                        elif file_format == 'xlsx':
                            data = self.excel_to_json(file)
                        else:
                            messages.error(request, "Неподдерживаемый формат файла.")
                            return self.import_products_view(request)

                        logger.info(f"Обработка {len(data)} записей из файла")
                        products_to_create = []
                        category_cache = {}
                        for index, row in enumerate(data):
                            try:
                                if not row.get('name'):
                                    errors.append(f"Запись {index + 1}: Поле 'name' обязательно.")
                                    continue
                                if 'price' not in row or row['price'] is None:
                                    errors.append(f"Запись {index + 1}: Поле 'price' обязательно.")
                                    continue
                                price = float(row['price'])
                                if price < 0:
                                    errors.append(f"Запись {index + 1}: Цена не может быть отрицательной.")
                                    continue
                                if not row.get('category_path'):
                                    errors.append(f"Запись {index + 1}: Поле 'category_path' обязательно.")
                                    continue

                                category = self.get_or_create_category(row['category_path'], category_cache)
                                if not category:
                                    errors.append(f"Запись {index + 1}: Не удалось создать категорию '{row['category_path']}'.")
                                    continue

                                photo_path = None
                                if row.get('photo_filename'):
                                    photo_path = os.path.join('product_photos', row['photo_filename'])
                                    if not default_storage.exists(photo_path):
                                        errors.append(f"Запись {index + 1}: Файл фотографии '{photo_path}' не найден.")
                                        continue

                                product = Product(
                                    name=row['name'],
                                    description=row.get('description', ''),
                                    price=price,
                                    category=category,
                                    is_active=row.get('is_active', True),
                                )
                                if photo_path:
                                    product.photo = photo_path
                                products_to_create.append(product)

                            except Exception as e:
                                errors.append(f"Запись {index + 1}: Ошибка: {str(e)}")

                        if products_to_create:
                            Product.objects.bulk_create(products_to_create)
                            imported_count = len(products_to_create)
                            logger.info(f"Создано {imported_count} товаров")

                    if imported_count > 0:
                        messages.success(request, f"Импортировано {imported_count} товаров.")
                    if errors:
                        for error in errors:
                            messages.error(request, error)
                    else:
                        messages.success(request, "Импорт завершён без ошибок.")

                except Exception as e:
                    logger.error(f"Ошибка при импорте: {str(e)}")
                    messages.error(request, f"Ошибка при импорте: {str(e)}")
                return redirect('admin:shop_product_changelist')
        else:
            form = ImportProductsForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Импорт товаров',
        }
        return render(request, 'admin/shop/product/import_products.html', context)

    def get_queryset(self, request):
        return Product.objects.filter(is_active=True)

    def save_model(self, request, obj, form, change):
        if change:
            logger.info(f'Товар изменён: {obj}')
        else:
            logger.info(f'Создан новый товар: {obj}')
        super().save_model(request, obj, form, change)

    def export_products(self, request, queryset):
        logger.info("Начало экспорта товаров")

        if 'export_form_submit' in request.POST:
            logger.info("Получен POST-запрос от формы экспорта")
            form = ProductExportForm(request.POST)
            if form.is_valid():
                logger.info("Форма валидна")
                file_format = form.cleaned_data['file_format']
                selected_fields = form.cleaned_data['fields']
                category = form.cleaned_data['category']
                is_active = form.cleaned_data['is_active']
                date_from = form.cleaned_data['date_from']
                date_to = form.cleaned_data['date_to']

                logger.info(f"Параметры формы: file_format={file_format}, selected_fields={selected_fields}, "
                           f"category={category}, is_active={is_active}, date_from={date_from}, date_to={date_to}")

                selected_ids = request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)
                if selected_ids:
                    products_to_export = Product.objects.filter(id__in=selected_ids)
                    logger.info(f"Выбрано {products_to_export.count()} товаров через selected_ids")
                else:
                    products_to_export = Product.objects.all()
                    logger.info(f"Всего товаров в базе: {products_to_export.count()}")
                    if category:
                        products_to_export = products_to_export.filter(category=category)
                        logger.info(f"После фильтра по категории: {products_to_export.count()}")
                    if is_active:
                        products_to_export = products_to_export.filter(is_active=(is_active == '1'))
                        logger.info(f"После фильтра по статусу: {products_to_export.count()}")
                    if date_from:
                        products_to_export = products_to_export.filter(created_at__gte=date_from)
                        logger.info(f"После фильтра по дате (с): {products_to_export.count()}")
                    if date_to:
                        products_to_export = products_to_export.filter(created_at__lte=date_to)
                        logger.info(f"После фильтра по дате (по): {products_to_export.count()}")

                if not products_to_export.exists():
                    logger.warning("Нет товаров для экспорта после применения фильтров")
                    messages.warning(request, "Нет товаров для экспорта. Проверьте фильтры или добавьте товары в базу.")
                    return redirect('admin:shop_product_changelist')

                logger.info("Формирование данных для экспорта")
                data = []
                for product in products_to_export:
                    row = {}
                    if 'id' in selected_fields:
                        row['id'] = product.id
                    if 'name' in selected_fields:
                        row['name'] = product.name
                    if 'description' in selected_fields:
                        row['description'] = product.description
                    if 'price' in selected_fields:
                        row['price'] = float(product.price)
                    if 'category_path' in selected_fields:
                        row['category_path'] = self.get_category_path(product.category)
                    if 'photo_filename' in selected_fields:
                        row['photo_filename'] = os.path.basename(product.photo.name) if product.photo else ''
                    if 'is_active' in selected_fields:
                        row['is_active'] = product.is_active
                    if 'created_at' in selected_fields:
                        row['created_at'] = product.created_at.isoformat()
                    data.append(row)

                logger.info(f"Сформировано {len(data)} записей для экспорта")

                try:
                    if file_format == 'json':
                        logger.info("Генерация JSON-файла")
                        response = HttpResponse(content_type='application/json')
                        response['Content-Disposition'] = 'attachment; filename="products_export.json"'
                        response.write(json.dumps(data, ensure_ascii=False, indent=2))
                        logger.info("Файл JSON сгенерирован и отправлен")
                    elif file_format == 'csv':
                        logger.info("Генерация CSV-файла")
                        response = HttpResponse(content_type='text/csv')
                        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
                        writer = csv.DictWriter(response, fieldnames=selected_fields)
                        writer.writeheader()
                        for row in data:
                            writer.writerow(row)
                        logger.info("Файл CSV сгенерирован и отправлен")
                    elif file_format == 'xlsx':
                        logger.info("Генерация XLSX-файла")
                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        sheet.append(selected_fields)
                        for row in data:
                            sheet.append([row.get(field, '') for field in selected_fields])
                        workbook.save(response)
                        logger.info("Файл Excel сгенерирован и отправлен")
                    else:
                        logger.error(f"Неизвестный формат файла: {file_format}")
                        messages.error(request, "Неизвестный формат файла.")
                        return redirect('admin:shop_product_changelist')

                    return response

                except Exception as e:
                    logger.error(f"Ошибка при генерации файла: {str(e)}")
                    messages.error(request, f"Ошибка при генерации файла: {str(e)}")
                    return redirect('admin:shop_product_changelist')

            else:
                logger.warning("Форма невалидна")
                logger.warning(form.errors)
                messages.error(request, "Форма заполнена некорректно. Проверьте введённые данные.")
                return render(request, 'admin/shop/order/export_products.html', {'form': form})

        if request.method == "POST" and 'action' in request.POST:
            logger.info("Получен POST-запрос от админки для действия export_products")
            selected_ids = request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)
            if not selected_ids:
                self.message_user(request, "Пожалуйста, выберите хотя бы один товар для экспорта.", level=messages.WARNING)
                return redirect('admin:shop_product_changelist')

            form = ProductExportForm()
            context = {
                'form': form,
                'queryset': Product.objects.filter(id__in=selected_ids),
                'action': 'export_products',
                'opts': self.model._meta,
                'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
            }
            return render(request, 'admin/shop/order/export_products.html', context)

        form = ProductExportForm()
        return render(request, 'admin/shop/order/export_products.html', {'form': form})

    export_products.short_description = "Экспортировать товары"

    def parse_json(self, file):
        content = file.read().decode('utf-8')
        return json.loads(content)

    def csv_to_json(self, file):
        content = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(content))
        return [dict(row) for row in reader]

    def excel_to_json(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header, value in zip(headers, row):
                row_data[header] = value
            data.append(row_data)
        return data

    def get_or_create_category(self, category_path, category_cache=None):
        if category_cache is None:
            category_cache = {}

        cache_key = category_path
        if cache_key in category_cache:
            return category_cache[cache_key]

        parts = category_path.split('/')
        parent = None
        current_path = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            current_path.append(part)
            path_key = '/'.join(current_path)
            if path_key in category_cache:
                parent = category_cache[path_key]
                continue
            category, created = Category.objects.get_or_create(
                name=part,
                parent=parent,
                defaults={'is_active': True}
            )
            category_cache[path_key] = category
            parent = category

        return parent

    def get_category_path(self, category):
        if not category:
            return ''
        ancestors = category.get_ancestors(include_self=True)
        return '/'.join(ancestor.name for ancestor in ancestors)

    def name_colored(self, obj):
        return super().name_colored(obj)
    name_colored.short_description = 'Название'
