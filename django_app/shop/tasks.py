import logging
import openpyxl
from django.core.files.storage import default_storage
from django.db import transaction
from .models import Order, OrderItem
import aiohttp
import asyncio
from django.conf import settings
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

# Настройка логирования для данного модуля
logger = logging.getLogger(__name__)

# Получаем токен бота из настроек
BOT_TOKEN = settings.BOT_TOKEN

async def send_telegram_message(chat_id, text):
    """
    Асинхронная функция для отправки сообщения в Telegram.

    Args:
        chat_id (int): ID чата Telegram (пользователя).
        text (str): Текст сообщения.

    Returns:
        bool: True, если сообщение отправлено успешно, False в противном случае.
    """
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN не указан в настройках. Уведомления не могут быть отправлены.")
        return False

    telegram_url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    params = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML"
    }

    async with aiohttp.ClientSession() as session:
        try:
            async with session.post(telegram_url, json=params) as response:
                if response.status == 200:
                    logger.info(f"Сообщение отправлено пользователю {chat_id}: {text}")
                    return True
                else:
                    error_data = await response.json()
                    error_description = error_data.get("description", "Неизвестная ошибка")
                    if error_data.get("error_code") == 403:
                        logger.warning(f"Не удалось отправить сообщение пользователю {chat_id}: Пользователь заблокировал бота или не начинал диалог. Описание: {error_description}")
                    elif error_data.get("error_code") == 404:
                        logger.error(f"Не удалось отправить сообщение пользователю {chat_id}: Неверный BOT_TOKEN или пользователь не найден. Описание: {error_description}")
                    else:
                        logger.error(f"Ошибка отправки сообщения пользователю {chat_id}: {error_description}")
                    return False
        except Exception as e:
            logger.error(f"Исключение при отправке сообщения пользователю {chat_id}: {e}")
            return False

def export_orders_to_excel(queryset=None):
    """
    Экспорт заказов в Excel-файл.

    Args:
        queryset: QuerySet заказов для экспорта. Если None, экспортируются все активные заказы.

    Создаёт Excel-файл с информацией о заказах, включая ID заказа, пользователя,
    адрес доставки, телефон, пожелания, желаемое время доставки, статус и список товаров.

    Возвращает путь к сохранённому Excel-файлу.
    """
    logger.info('Начало экспорта заказов в Excel.')

    try:
        # Создание новой рабочей книги и настройка листа
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Заказы"
        logger.debug('Создана новая рабочая книга Excel.')

        # Установка заголовков столбцов
        headers = [
            "ID заказа",
            "Пользователь",
            "Адрес доставки",
            "Телефон",
            "Пожелания",
            "Желаемое время доставки",
            "Статус",
            "Товары"
        ]
        sheet.append(headers)
        logger.debug('Установлены заголовки столбцов в Excel.')

        # Инициализация номера строки для заполнения данных
        row_number = 2

        # Получение заказов для экспорта
        if queryset is None:
            orders = Order.objects.filter(is_active=True)
        else:
            orders = queryset
        logger.info(f'Получено {orders.count()} заказов для экспорта.')

        for order in orders:
            # Получение товаров, связанных с заказом
            order_items = OrderItem.objects.filter(order=order, is_active=True)
            logger.debug(f'Получено {order_items.count()} товаров для заказа №{order.id}.')

            items_list = [f"{item.product.name} x {item.quantity}" for item in order_items]

            # Заполнение строки данными заказа
            sheet.cell(row=row_number, column=1, value=order.id)
            sheet.cell(row=row_number, column=2, value=order.user.username or f"User {order.user.telegram_id}")
            sheet.cell(row=row_number, column=3, value=order.address)
            sheet.cell(row=row_number, column=4, value=order.phone or "-")
            sheet.cell(row=row_number, column=5, value=order.wishes or "-")
            sheet.cell(row=row_number, column=6, value=order.desired_delivery_time or "-")
            sheet.cell(row=row_number, column=7, value=order.get_status_display())
            sheet.cell(row=row_number, column=8, value=", ".join(items_list) if items_list else "Нет товаров")

            logger.debug(f'Заполнена строка для заказа №{order.id}.')
            row_number += 1

        # Определение имени файла и сохранение рабочей книги
        excel_filename = "orders_export.xlsx"
        workbook_path = default_storage.path(excel_filename)
        workbook.save(workbook_path)
        logger.info(f'Экспорт заказов завершён. Файл сохранён по пути: {workbook_path}')

        return workbook_path

    except Exception as e:
        logger.error(f"Ошибка при экспорте заказов в Excel: {e}", exc_info=True)
        return None

def notify_user_of_status_change(order_id, old_status, new_status):
    try:
        from .models import Order  # Локальный импорт для избежания циклических зависимостей
        
        order = Order.objects.get(id=order_id)
        user = order.user
        
        if not user.telegram_id:
            logger.warning(f"У пользователя нет telegram_id: {user}")
            return
            
        status_names = dict(Order.STATUS_CHOICES)
        message = (
            f"🔄 Статус вашего заказа №{order_id} изменён:\n"
            f"Было: {status_names.get(old_status, old_status)}\n"
            f"Стало: {status_names.get(new_status, new_status)}"
        )
        
        # Используем синхронную версию отправки
        send_telegram_message_sync(user.telegram_id, message)
        
    except Exception as e:
        logger.error(f"Ошибка уведомления: {e}")


def send_telegram_message_sync(chat_id, text):
    """Синхронная версия отправки сообщения"""
    import requests
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    try:
        response = requests.post(url, json={
            "chat_id": chat_id,
            "text": text,
            "parse_mode": "HTML"
        })
        return response.status_code == 200
    except Exception as e:
        logger.error(f"Ошибка отправки: {e}")
